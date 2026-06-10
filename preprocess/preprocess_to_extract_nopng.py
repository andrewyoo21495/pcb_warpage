"""
PCB Warpage: Preprocessing + Feature Extraction Pipeline (No PNG, Standalone)
==============================================================================
Self-contained pipeline from raw elevation data to feature Excel output.
No external local module imports — all functions are embedded in this file.

Steps:
  1. Preprocess (downsample, tilt correction, outlier removal, interpolation, smoothing)
  2. Save preprocessed .txt only (no grayscale image)
  3. Extract 12 warpage feature variables
  4. Export features to Excel

Requirements:
    pip install numpy scipy scikit-learn openpyxl

Usage:
    # Single file
    python preprocess_to_extract_nopng.py --input-file /path/to/raw.txt

    # Directory (all .txt files in folder)
    python preprocess_to_extract_nopng.py --input-dir /path/to/folder

    # With custom preprocessing parameters
    python preprocess_to_extract_nopng.py --input-dir /path/to/folder \
        --downsample-factor 4 --z-threshold 3.0 --poly-degree 3

    # Custom output Excel path
    python preprocess_to_extract_nopng.py --input-dir /path/to/folder \
        --output features.xlsx
"""

import argparse
import logging
import os
import sys
import time
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
from scipy.ndimage import gaussian_filter
from scipy.stats import skew, kurtosis
from sklearn.linear_model import Ridge
from sklearn.preprocessing import PolynomialFeatures

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("[ERROR] openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

logger = logging.getLogger(__name__)


# =============================================================================
# Configuration
# =============================================================================

EXCLUDE_SUFFIXES = ("ORI", "ORI@LOW", "ORI_A")


@dataclass
class PreprocessorConfig:
    """Configuration for the PCB elevation preprocessing pipeline."""

    # File filtering
    exclude_suffixes: List[str] = field(
        default_factory=lambda: ["ORI", "ORI@LOW", "ORI_A"]
    )

    # Null value sentinel
    null_value: float = 9999.0

    # Downsampling
    downsample_factor: int = 4

    # Outlier detection
    outlier_z_threshold: float = 3.0
    outlier_grid_size: int = 8

    # Polynomial interpolation
    interp_poly_degree: int = 3
    interp_ridge_alpha: float = 0.1

    # Tilt correction
    tilt_patch_size: int = 16

    # Gaussian smoothing
    gaussian_sigma: float = 2.0
    gaussian_iterations: int = 3

    # Fixed scaling range (kept for compatibility, not used for imaging here)
    scale_min: float = 0.0
    scale_max: float = 3000.0


# =============================================================================
# File I/O Utilities (from preprocess_total.py)
# =============================================================================

def should_skip_file(filepath: str, exclude_suffixes: list = EXCLUDE_SUFFIXES) -> bool:
    """Returns True if the filename (without extension) ends with any of the exclude_suffixes."""
    stem = Path(filepath).stem
    return any(stem.endswith(suffix) for suffix in exclude_suffixes)


def read_elevation(filepath: str, null_value: float = 9999.0) -> np.ndarray:
    """Load a tab-delimited elevation file and replace null sentinels with NaN."""
    data = np.loadtxt(filepath, delimiter='\t')
    data[data == null_value] = np.nan
    return data


def save_preprocessed_txt(data: np.ndarray, output_path: str) -> None:
    """Save preprocessed data as tab-delimited text with 4 decimal places."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    np.savetxt(output_path, data, delimiter='\t', fmt='%.4f')


# =============================================================================
# Preprocessing Functions (from preprocess_total.py)
# =============================================================================

def downsample_median(data: np.ndarray, factor: int) -> np.ndarray:
    """Downsample data by computing the median of non-NaN values in each block.

    Args:
        data: Input array of shape (H, W), may contain NaN.
        factor: Downsampling factor (e.g., 4 means 1/4 resolution).

    Returns:
        Downsampled array of shape (H // factor, W // factor).
    """
    H, W = data.shape
    new_H = H // factor
    new_W = W // factor

    # Truncate to exact multiple of factor
    trimmed = data[:new_H * factor, :new_W * factor]

    # Vectorized: reshape into blocks and compute nanmedian in one call
    reshaped = trimmed.reshape(new_H, factor, new_W, factor)
    reshaped = reshaped.transpose(0, 2, 1, 3).reshape(new_H, new_W, factor * factor)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", RuntimeWarning)
        result = np.nanmedian(reshaped, axis=2)

    return result


def detect_and_remove_outliers(
    data: np.ndarray,
    grid_size: int = 8,
    z_threshold: float = 3.0,
) -> tuple:
    """Detect outliers per region using z-score and replace with NaN.

    Returns:
        (result_array, total_outliers_removed)
    """
    result = data.copy()
    H, W = result.shape
    total_removed = 0

    row_edges = np.linspace(0, H, grid_size + 1, dtype=int)
    col_edges = np.linspace(0, W, grid_size + 1, dtype=int)

    for ri in range(grid_size):
        for ci in range(grid_size):
            r_start, r_end = row_edges[ri], row_edges[ri + 1]
            c_start, c_end = col_edges[ci], col_edges[ci + 1]
            region = result[r_start:r_end, c_start:c_end]

            valid_mask = ~np.isnan(region)
            valid_vals = region[valid_mask]

            if len(valid_vals) < 2:
                continue

            mean = np.mean(valid_vals)
            std = np.std(valid_vals, ddof=0)

            if std == 0:
                continue

            z_scores = np.abs((region - mean) / std)
            outlier_mask = valid_mask & (z_scores > z_threshold)
            n_outliers = np.sum(outlier_mask)

            if n_outliers > 0:
                region[outlier_mask] = np.nan
                total_removed += n_outliers

    return result, int(total_removed)


def interpolate_surface(
    data: np.ndarray,
    poly_degree: int = 3,
    ridge_alpha: float = 0.1,
) -> tuple:
    """Fill NaN values using polynomial surface regression with ridge regularization.

    Returns:
        (result_array, n_interpolated) — array with NaNs filled, and count of filled pixels.
    """
    H, W = data.shape
    total_pixels = H * W

    rows, cols = np.indices((H, W))
    rows_flat = rows.ravel().astype(np.float64)
    cols_flat = cols.ravel().astype(np.float64)
    vals_flat = data.ravel()

    valid_mask = ~np.isnan(vals_flat)
    nan_mask = np.isnan(vals_flat)

    n_valid = np.sum(valid_mask)
    n_interpolated = int(np.sum(nan_mask))
    valid_ratio = n_valid / total_pixels

    if n_valid == 0:
        logger.warning("No valid values — cannot interpolate.")
        return data, 0

    if valid_ratio < 0.05:
        logger.warning("Valid values: %.1f%% (< 5%%) — potential quality degradation.",
                        valid_ratio * 100)

    # Normalize coordinates for numerical stability
    row_norm = rows_flat / max(H - 1, 1)
    col_norm = cols_flat / max(W - 1, 1)
    coords = np.column_stack([row_norm, col_norm])

    # Polynomial features
    poly = PolynomialFeatures(degree=poly_degree, include_bias=True)
    X_all = poly.fit_transform(coords)

    X_train = X_all[valid_mask]
    y_train = vals_flat[valid_mask]

    # Fit ridge regression
    model = Ridge(alpha=ridge_alpha)
    model.fit(X_train, y_train)

    # Predict NaN positions only
    result = data.copy()
    if np.any(nan_mask):
        X_predict = X_all[nan_mask]
        predicted = model.predict(X_predict)
        result_flat = result.ravel()
        result_flat[nan_mask] = predicted
        result = result_flat.reshape(H, W)

    return result, n_interpolated


def flatten_tilt(data: np.ndarray, patch_size: int = 16) -> tuple:
    """Remove linear tilt by fitting and subtracting a plane through corner patches.

    Computes the mean elevation of each corner patch (using nanmean to handle
    NaN values), fits a least-squares plane z = a*x + b*y + c through the
    valid corner points, and subtracts it from the surface.  The result is
    shifted so that its minimum is zero.

    Args:
        data: Input 2D array (H, W), may contain NaN values.
        patch_size: Side length of the square patch at each corner used to
                    compute stable corner elevation estimates.

    Returns:
        (flattened_data, plane_amplitude) where plane_amplitude is the max-min
        of the subtracted plane (a diagnostic for how much tilt was removed).
    """
    H, W = data.shape
    ps = min(patch_size, H // 4, W // 4)  # clamp to avoid overlap

    # Corner patches: mean elevation and centroid coordinates (row, col)
    corners = [
        (data[:ps, :ps],             ps / 2,       ps / 2),        # top-left
        (data[:ps, W - ps:],         ps / 2,       W - ps / 2),    # top-right
        (data[H - ps:, :ps],         H - ps / 2,   ps / 2),        # bottom-left
        (data[H - ps:, W - ps:],     H - ps / 2,   W - ps / 2),   # bottom-right
    ]

    # Build system using only corners with valid (non-NaN) data
    A_rows = []
    z_vals = []
    for patch, r_center, c_center in corners:
        patch_mean = float(np.nanmean(patch))
        if np.isnan(patch_mean):
            continue
        A_rows.append([r_center, c_center, 1.0])
        z_vals.append(patch_mean)

    n_valid = len(z_vals)
    if n_valid < 3:
        logger.warning("Only %d valid corner patches — skipping tilt correction.", n_valid)
        return data.copy().astype(np.float32), 0.0

    A = np.array(A_rows, dtype=np.float64)
    z = np.array(z_vals, dtype=np.float64)

    # Least-squares solve (exact for 3 points, overdetermined for 4)
    coeffs, _, _, _ = np.linalg.lstsq(A, z, rcond=None)  # [a, b, c]

    # Build the full plane surface
    rows, cols = np.indices((H, W), dtype=np.float64)
    plane = coeffs[0] * rows + coeffs[1] * cols + coeffs[2]

    plane_amplitude = float(plane.max() - plane.min())

    # Subtract plane and shift minimum to zero (using nanmin to handle NaN)
    flattened = data.astype(np.float64) - plane
    flattened -= np.nanmin(flattened)

    return flattened.astype(np.float32), plane_amplitude


def smooth_gaussian(
    data: np.ndarray,
    sigma: float = 2.0,
    iterations: int = 3,
) -> np.ndarray:
    """Apply iterative Gaussian smoothing while preserving the original min/max.

    Sigma is adaptive: scaled proportionally to the data dimensions so that
    the visual smoothness is consistent regardless of resolution.
    """
    rows, cols = data.shape
    sigma_row = max(1.0, rows * sigma / 100)
    sigma_col = max(1.0, cols * sigma / 100)

    orig_min = np.min(data)
    orig_max = np.max(data)

    if orig_max - orig_min < 1e-12:
        return data.copy()

    smoothed = data.copy()
    for _ in range(iterations):
        smoothed = gaussian_filter(smoothed, sigma=[sigma_row, sigma_col])

        s_min = np.min(smoothed)
        s_max = np.max(smoothed)

        if s_max - s_min < 1e-12:
            break

        smoothed = (smoothed - s_min) / (s_max - s_min)
        smoothed = smoothed * (orig_max - orig_min) + orig_min

    return smoothed


# =============================================================================
# Preprocessing Orchestration (from preprocess_total.py)
# =============================================================================

def process_single_file(filepath: str, config: PreprocessorConfig) -> tuple:
    """Run Steps 1-6 on a single file and return the preprocessed data with stats.

    Returns:
        (data, n_outliers, n_interpolated, plane_amplitude) or
        (None, 0, 0, 0.0) if the file should be skipped.
    """
    try:
        data = read_elevation(filepath, null_value=config.null_value)
    except Exception as e:
        logger.warning("Failed to read %s: %s", filepath, e)
        return None, 0, 0, 0.0

    if np.all(np.isnan(data)):
        return None, 0, 0, 0.0

    data = downsample_median(data, factor=config.downsample_factor)
    data, plane_amplitude = flatten_tilt(
        data, patch_size=config.tilt_patch_size,
    )
    data, n_outliers = detect_and_remove_outliers(
        data,
        grid_size=config.outlier_grid_size,
        z_threshold=config.outlier_z_threshold,
    )
    data, n_interpolated = interpolate_surface(
        data,
        poly_degree=config.interp_poly_degree,
        ridge_alpha=config.interp_ridge_alpha,
    )
    data = smooth_gaussian(
        data, sigma=config.gaussian_sigma, iterations=config.gaussian_iterations,
    )

    return data, n_outliers, n_interpolated, plane_amplitude


# =============================================================================
# Feature Extraction Functions (from wap_parameter_extract.py)
# =============================================================================

def extract_basic_statistics(Z: np.ndarray) -> Dict[str, float]:
    """Extract basic statistical features from the warpage surface.

    Variables:
        max_warpage, skewness, kurtosis
    """
    vals = Z.ravel()
    return {
        "max_warpage": float(np.max(vals)),
        "skewness": float(skew(vals)),
        "kurtosis": float(kurtosis(vals)),  # excess kurtosis (Fisher)
    }


def extract_warpage_mode(Z: np.ndarray) -> Dict[str, float]:
    """Extract warpage mode classification features.

    Variables:
        center_vs_edge_ratio, saddle_index, corner_elevation_std,
        warpage_mode, edge_profile_asymmetry
    """
    H, W = Z.shape

    # --- center_vs_edge_ratio ---
    r_start, r_end = H // 4, 3 * H // 4
    c_start, c_end = W // 4, 3 * W // 4
    center_mean = float(np.mean(Z[r_start:r_end, c_start:c_end]))

    edge_mask = np.ones_like(Z, dtype=bool)
    edge_mask[r_start:r_end, c_start:c_end] = False
    edge_mean = float(np.mean(Z[edge_mask]))

    center_vs_edge_ratio = center_mean / edge_mean if edge_mean != 0 else float('inf')

    # --- Corner elevations (using corner patches) ---
    ps = min(H // 8, W // 8, 16)
    ps = max(ps, 1)
    corner_tl = float(np.mean(Z[:ps, :ps]))
    corner_tr = float(np.mean(Z[:ps, W - ps:]))
    corner_bl = float(np.mean(Z[H - ps:, :ps]))
    corner_br = float(np.mean(Z[H - ps:, W - ps:]))

    # --- saddle_index ---
    saddle_index = (corner_tl + corner_br) / 2 - (corner_tr + corner_bl) / 2

    # --- corner_elevation_std ---
    corners = [corner_tl, corner_tr, corner_bl, corner_br]
    corner_elevation_std = float(np.std(corners))

    # --- edge_profile_asymmetry ---
    edge_top = float(np.mean(Z[:ps, :]))
    edge_bottom = float(np.mean(Z[H - ps:, :]))
    edge_left = float(np.mean(Z[:, :ps]))
    edge_right = float(np.mean(Z[:, W - ps:]))
    edge_profile_asymmetry = float(np.std([edge_top, edge_bottom, edge_left, edge_right]))

    # --- warpage_mode classification ---
    z_range = float(np.max(Z) - np.min(Z))
    saddle_threshold = z_range * 0.1
    corner_std_threshold = z_range * 0.1

    if corner_elevation_std > corner_std_threshold and abs(saddle_index) > saddle_threshold:
        if abs(saddle_index) > corner_elevation_std:
            warpage_mode = "Saddle"
        else:
            warpage_mode = "Twist"
    elif center_vs_edge_ratio > 1.05:
        warpage_mode = "Crying"
    elif center_vs_edge_ratio < 0.95:
        warpage_mode = "Smiling"
    else:
        warpage_mode = "Flat"

    return {
        "center_vs_edge_ratio": round(center_vs_edge_ratio, 6),
        "saddle_index": round(saddle_index, 4),
        "corner_elevation_std": round(corner_elevation_std, 4),
        "warpage_mode": warpage_mode,
        "edge_profile_asymmetry": round(edge_profile_asymmetry, 4),
    }


def extract_surface_geometry(Z: np.ndarray) -> Dict[str, float]:
    """Extract surface geometry features (gradient, curvature).

    Variables:
        mean_gradient, max_gradient, mean_laplacian, laplacian_std
    """
    gy, gx = np.gradient(Z)
    grad_mag = np.sqrt(gx**2 + gy**2)

    lap_yy = np.gradient(np.gradient(Z, axis=0), axis=0)
    lap_xx = np.gradient(np.gradient(Z, axis=1), axis=1)
    laplacian = lap_yy + lap_xx

    return {
        "mean_gradient": round(float(np.mean(grad_mag)), 4),
        "max_gradient": round(float(np.max(grad_mag)), 4),
        "mean_laplacian": round(float(np.mean(laplacian)), 4),
        "laplacian_std": round(float(np.std(laplacian)), 4),
    }


def extract_all_features(Z: np.ndarray) -> Dict[str, object]:
    """Extract all selected features from a warpage surface."""
    features = {}
    features.update(extract_basic_statistics(Z))
    features.update(extract_warpage_mode(Z))
    features.update(extract_surface_geometry(Z))
    return features


# Column order for output
FEATURE_COLUMNS = [
    # Basic statistics
    "max_warpage", "skewness", "kurtosis",
    # Warpage mode
    "center_vs_edge_ratio", "saddle_index", "corner_elevation_std",
    "warpage_mode", "edge_profile_asymmetry",
    # Surface geometry
    "mean_gradient", "max_gradient", "mean_laplacian", "laplacian_std",
]


# =============================================================================
# Excel Export (from wap_parameter_extract.py)
# =============================================================================

CATEGORY_COLORS = {
    "Basic Statistics": {
        "fill": "E8F0FE",
        "columns": ["max_warpage", "skewness", "kurtosis"],
    },
    "Warpage Mode": {
        "fill": "FCE8E6",
        "columns": ["center_vs_edge_ratio", "saddle_index",
                     "corner_elevation_std", "warpage_mode",
                     "edge_profile_asymmetry"],
    },
    "Surface Geometry": {
        "fill": "FEF7E0",
        "columns": ["mean_gradient", "max_gradient",
                     "mean_laplacian", "laplacian_std"],
    },
}


def export_to_excel(results: List[Dict], output_path: str) -> None:
    """Export extraction results to a styled Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Warpage Features"

    # --- Styles ---
    header_font = Font(bold=True, size=10)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    pcb_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # --- Column header row (row 1) ---
    all_columns = ["pcb_name"] + FEATURE_COLUMNS
    for ci, col_name in enumerate(all_columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        if col_name == "pcb_name":
            cell.fill = pcb_fill

    # --- Data rows (row 2+) ---
    for ri, record in enumerate(results, start=2):
        for ci, col_name in enumerate(all_columns, start=1):
            value = record.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 30  # pcb_name
    for ci in range(2, len(all_columns) + 1):
        ws.column_dimensions[chr(64 + ci) if ci <= 26
                             else chr(64 + (ci - 1) // 26) + chr(64 + (ci - 1) % 26 + 1)].width = 18

    # Freeze panes: fix header row and pcb_name column
    ws.freeze_panes = "B2"

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    wb.save(output_path)


# =============================================================================
# File Discovery
# =============================================================================

def discover_txt_files(directory: str, exclude_suffixes: list) -> list:
    """Find all .txt files in a directory, excluding skip-suffixes and output dirs."""
    txt_files = []
    for entry in sorted(os.listdir(directory)):
        if not entry.lower().endswith('.txt'):
            continue
        full_path = os.path.join(directory, entry)
        if not os.path.isfile(full_path):
            continue
        if should_skip_file(full_path, exclude_suffixes):
            continue
        txt_files.append(full_path)
    return txt_files


# =============================================================================
# Pipeline
# =============================================================================

def run_pipeline(filepaths: list, config: PreprocessorConfig,
                 output_xlsx: str) -> None:
    """Run preprocessing + feature extraction on a list of files."""
    total = len(filepaths)
    results = []
    files_ok = 0
    files_fail = 0

    print(f"\n  Processing {total} file(s)...\n")
    print(f"  {'#':>4s}  {'File':<40s}  {'Status':<8s}  {'Max':>8s}  {'Mode':<8s}")
    print(f"  {'':->4s}  {'':->40s}  {'':->8s}  {'':->8s}  {'':->8s}")

    for i, filepath in enumerate(filepaths, start=1):
        stem = Path(filepath).stem
        parent_dir = os.path.dirname(filepath)
        display_name = stem if len(stem) <= 40 else stem[:37] + "..."

        # --- Phase 1: Preprocess ---
        data, n_outliers, n_interp, plane_amp = process_single_file(filepath, config)

        if data is None:
            print(f"  {i:4d}  {display_name:<40s}  SKIP")
            files_fail += 1
            continue

        # Save preprocessed txt
        txt_out = os.path.join(parent_dir, "interpolated", f"{stem}_preprocessed.txt")
        save_preprocessed_txt(data, txt_out)

        # --- Phase 2: Extract features ---
        features = extract_all_features(data)
        record = {"pcb_name": stem}
        record.update(features)
        results.append(record)

        mode = features.get("warpage_mode", "?")
        max_w = features.get("max_warpage", 0)
        print(f"  {i:4d}  {display_name:<40s}  OK        {max_w:8.1f}  {mode:<8s}")
        files_ok += 1

    # --- Summary ---
    print(f"\n  {'':->70s}")
    print(f"  Preprocessed: {files_ok} OK, {files_fail} skipped")

    if not results:
        print("  No files were successfully processed. No Excel output.")
        return

    # --- Phase 3: Export to Excel ---
    print(f"  Exporting {len(results)} records to: {output_xlsx}")
    export_to_excel(results, output_xlsx)
    print(f"  Excel saved: {output_xlsx}")


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="PCB Warpage: Preprocessing + Feature Extraction Pipeline (No PNG, Standalone)"
    )

    # Input mode
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--input-file", type=str,
                       help="Single raw .txt file to process.")
    group.add_argument("--input-dir", type=str,
                       help="Directory containing raw .txt files.")

    # Preprocessing parameters
    parser.add_argument("--downsample-factor", type=int, default=4)
    parser.add_argument("--z-threshold", type=float, default=3.0)
    parser.add_argument("--grid-size", type=int, default=8)
    parser.add_argument("--poly-degree", type=int, default=3)
    parser.add_argument("--ridge-alpha", type=float, default=0.1)
    parser.add_argument("--tilt-patch-size", type=int, default=16)
    parser.add_argument("--gaussian-sigma", type=float, default=2.0)
    parser.add_argument("--smooth-iterations", type=int, default=3)

    # Output
    parser.add_argument("--output", type=str, default=None,
                        help="Output Excel file path (default: warpage_features.xlsx in input location).")

    args = parser.parse_args()

    # Build config
    config = PreprocessorConfig(
        downsample_factor=args.downsample_factor,
        outlier_z_threshold=args.z_threshold,
        outlier_grid_size=args.grid_size,
        interp_poly_degree=args.poly_degree,
        interp_ridge_alpha=args.ridge_alpha,
        tilt_patch_size=args.tilt_patch_size,
        gaussian_sigma=args.gaussian_sigma,
        gaussian_iterations=args.smooth_iterations,
    )

    start_time = time.time()
    print("\n" + "=" * 70)
    print("  PCB Warpage: Preprocessing + Feature Extraction Pipeline (No PNG)")
    print("=" * 70)

    # Discover files
    if args.input_file:
        if not os.path.isfile(args.input_file):
            print(f"  [ERROR] File not found: {args.input_file}")
            sys.exit(1)
        filepaths = [args.input_file]
        default_output_dir = os.path.dirname(args.input_file) or "."
        print(f"  Mode: Single file")
        print(f"  Input: {args.input_file}")
    else:
        if not os.path.isdir(args.input_dir):
            print(f"  [ERROR] Directory not found: {args.input_dir}")
            sys.exit(1)
        filepaths = discover_txt_files(args.input_dir, config.exclude_suffixes)
        default_output_dir = args.input_dir
        print(f"  Mode: Directory ({len(filepaths)} .txt files found)")
        print(f"  Input: {args.input_dir}")

    if not filepaths:
        print("  No .txt files found. Nothing to do.")
        sys.exit(0)

    output_xlsx = args.output or os.path.join(default_output_dir, "warpage_features.xlsx")
    print(f"  Output: {output_xlsx}")

    # Run pipeline
    run_pipeline(filepaths, config, output_xlsx)

    elapsed = time.time() - start_time
    print(f"\n{'=' * 70}")
    print(f"  Completed in {elapsed:.1f} seconds.")
    print(f"{'=' * 70}\n")


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.WARNING,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    try:
        main()
    except Exception as e:
        print(f"\n  [ERROR] Pipeline failed: {e}", file=sys.stderr)
        sys.exit(1)
