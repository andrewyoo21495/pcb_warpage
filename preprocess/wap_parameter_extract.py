"""
PCB Warpage Parameter Extractor
================================
Extracts warpage feature variables from preprocessed PCB elevation data
and exports them as an Excel file.

Requirements:
    pip install numpy scipy openpyxl

Usage:
    # Single file mode
    python preprocess/wap_parameter_extract.py --input-file /path/to/sample.txt

    # Folder mode (all .txt files in folder)
    python preprocess/wap_parameter_extract.py --input-dir /path/to/folder

    # Custom output path
    python preprocess/wap_parameter_extract.py --input-dir /path/to/folder --output results.xlsx
"""

import argparse
import os
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
from scipy.stats import skew, kurtosis

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("[ERROR] openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# Feature Extraction Functions
# =============================================================================

def extract_basic_statistics(Z: np.ndarray) -> Dict[str, float]:
    """Extract basic statistical features from the warpage surface.

    Variables:
        max_warpage, skewness, kurtosis
    """
    vals = Z.ravel()
    return {
        "max_warpage": float(np.max(vals)),
        "skewness": float(skew(vals)),
        "kurtosis": float(kurtosis(vals)),  # excess kurtosis (Fisher)
    }


def extract_warpage_mode(Z: np.ndarray) -> Dict[str, float]:
    """Extract warpage mode classification features.

    Variables:
        center_vs_edge_ratio, saddle_index, corner_elevation_std,
        warpage_mode, edge_profile_asymmetry
    """
    H, W = Z.shape

    # --- center_vs_edge_ratio ---
    # Center region: middle 50% in both dimensions (= 25% of total area)
    r_start, r_end = H // 4, 3 * H // 4
    c_start, c_end = W // 4, 3 * W // 4
    center_mean = float(np.mean(Z[r_start:r_end, c_start:c_end]))

    # Edge region: all pixels NOT in center
    edge_mask = np.ones_like(Z, dtype=bool)
    edge_mask[r_start:r_end, c_start:c_end] = False
    edge_mean = float(np.mean(Z[edge_mask]))

    center_vs_edge_ratio = center_mean / edge_mean if edge_mean != 0 else float('inf')

    # --- Corner elevations (using corner patches) ---
    ps = min(H // 8, W // 8, 16)  # patch size
    ps = max(ps, 1)
    corner_tl = float(np.mean(Z[:ps, :ps]))
    corner_tr = float(np.mean(Z[:ps, W - ps:]))
    corner_bl = float(np.mean(Z[H - ps:, :ps]))
    corner_br = float(np.mean(Z[H - ps:, W - ps:]))

    # --- saddle_index ---
    saddle_index = (corner_tl + corner_br) / 2 - (corner_tr + corner_bl) / 2

    # --- corner_elevation_std ---
    corners = [corner_tl, corner_tr, corner_bl, corner_br]
    corner_elevation_std = float(np.std(corners))

    # --- edge_profile_asymmetry ---
    edge_top = float(np.mean(Z[:ps, :]))
    edge_bottom = float(np.mean(Z[H - ps:, :]))
    edge_left = float(np.mean(Z[:, :ps]))
    edge_right = float(np.mean(Z[:, W - ps:]))
    edge_profile_asymmetry = float(np.std([edge_top, edge_bottom, edge_left, edge_right]))

    # --- warpage_mode classification ---
    z_range = float(np.max(Z) - np.min(Z))
    saddle_threshold = z_range * 0.1
    corner_std_threshold = z_range * 0.1

    if corner_elevation_std > corner_std_threshold and abs(saddle_index) > saddle_threshold:
        if abs(saddle_index) > corner_elevation_std:
            warpage_mode = "Saddle"
        else:
            warpage_mode = "Twist"
    elif center_vs_edge_ratio > 1.05:
        warpage_mode = "Crying"
    elif center_vs_edge_ratio < 0.95:
        warpage_mode = "Smiling"
    else:
        warpage_mode = "Flat"

    return {
        "center_vs_edge_ratio": round(center_vs_edge_ratio, 6),
        "saddle_index": round(saddle_index, 4),
        "corner_elevation_std": round(corner_elevation_std, 4),
        "warpage_mode": warpage_mode,
        "edge_profile_asymmetry": round(edge_profile_asymmetry, 4),
    }


def extract_surface_geometry(Z: np.ndarray) -> Dict[str, float]:
    """Extract surface geometry features (gradient, curvature).

    Variables:
        mean_gradient, max_gradient, mean_laplacian, laplacian_std
    """
    # Gradient magnitude
    gy, gx = np.gradient(Z)
    grad_mag = np.sqrt(gx**2 + gy**2)

    # Laplacian (sum of second partial derivatives)
    lap_yy = np.gradient(np.gradient(Z, axis=0), axis=0)
    lap_xx = np.gradient(np.gradient(Z, axis=1), axis=1)
    laplacian = lap_yy + lap_xx

    return {
        "mean_gradient": round(float(np.mean(grad_mag)), 4),
        "max_gradient": round(float(np.max(grad_mag)), 4),
        "mean_laplacian": round(float(np.mean(laplacian)), 4),
        "laplacian_std": round(float(np.std(laplacian)), 4),
    }


def extract_all_features(Z: np.ndarray) -> Dict[str, object]:
    """Extract all selected features from a warpage surface."""
    features = {}
    features.update(extract_basic_statistics(Z))
    features.update(extract_warpage_mode(Z))
    features.update(extract_surface_geometry(Z))
    return features


# Column order for output
FEATURE_COLUMNS = [
    # Basic statistics
    "max_warpage", "skewness", "kurtosis",
    # Warpage mode
    "center_vs_edge_ratio", "saddle_index", "corner_elevation_std",
    "warpage_mode", "edge_profile_asymmetry",
    # Surface geometry
    "mean_gradient", "max_gradient", "mean_laplacian", "laplacian_std",
]


# =============================================================================
# File I/O
# =============================================================================

def read_elevation(filepath: str) -> np.ndarray:
    """Load a tab-delimited elevation file."""
    return np.loadtxt(filepath, delimiter='\t')


def discover_txt_files(directory: str) -> List[str]:
    """Find all .txt files in a directory (non-recursive)."""
    txt_files = []
    for entry in sorted(os.listdir(directory)):
        if entry.lower().endswith('.txt'):
            full_path = os.path.join(directory, entry)
            if os.path.isfile(full_path):
                txt_files.append(full_path)
    return txt_files


# =============================================================================
# Excel Export
# =============================================================================

# Category info for header coloring
CATEGORY_COLORS = {
    "Basic Statistics": {
        "fill": "E8F0FE",
        "columns": ["max_warpage", "skewness", "kurtosis"],
    },
    "Warpage Mode": {
        "fill": "FCE8E6",
        "columns": ["center_vs_edge_ratio", "saddle_index",
                     "corner_elevation_std", "warpage_mode",
                     "edge_profile_asymmetry"],
    },
    "Surface Geometry": {
        "fill": "FEF7E0",
        "columns": ["mean_gradient", "max_gradient",
                     "mean_laplacian", "laplacian_std"],
    },
}


def export_to_excel(results: List[Dict], output_path: str) -> None:
    """Export extraction results to a styled Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Warpage Features"

    # --- Styles ---
    header_font = Font(bold=True, size=10)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    pcb_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Build column-to-category mapping
    col_to_fill = {}
    for cat_info in CATEGORY_COLORS.values():
        fill = PatternFill(start_color=cat_info["fill"],
                           end_color=cat_info["fill"], fill_type="solid")
        for col_name in cat_info["columns"]:
            col_to_fill[col_name] = fill

    # --- Category header row (row 1) ---
    ws.cell(row=1, column=1, value="")
    col_idx = 2
    for cat_name, cat_info in CATEGORY_COLORS.items():
        n_cols = len(cat_info["columns"])
        start_col = col_idx
        end_col = col_idx + n_cols - 1
        cell = ws.cell(row=1, column=start_col, value=cat_name)
        cell.font = Font(bold=True, size=11)
        cell.alignment = header_alignment
        fill = PatternFill(start_color=cat_info["fill"],
                           end_color=cat_info["fill"], fill_type="solid")
        cell.fill = fill
        cell.border = thin_border
        if n_cols > 1:
            ws.merge_cells(start_row=1, start_column=start_col,
                           end_row=1, end_column=end_col)
            for c in range(start_col + 1, end_col + 1):
                ws.cell(row=1, column=c).fill = fill
                ws.cell(row=1, column=c).border = thin_border
        col_idx = end_col + 1

    # --- Column header row (row 2) ---
    all_columns = ["pcb_name"] + FEATURE_COLUMNS
    for ci, col_name in enumerate(all_columns, start=1):
        cell = ws.cell(row=2, column=ci, value=col_name)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        if col_name == "pcb_name":
            cell.fill = pcb_fill
        elif col_name in col_to_fill:
            cell.fill = col_to_fill[col_name]

    # --- Data rows (row 3+) ---
    for ri, record in enumerate(results, start=3):
        for ci, col_name in enumerate(all_columns, start=1):
            value = record.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 30  # pcb_name
    for ci in range(2, len(all_columns) + 1):
        ws.column_dimensions[chr(64 + ci) if ci <= 26
                             else chr(64 + (ci - 1) // 26) + chr(64 + (ci - 1) % 26 + 1)].width = 18

    # Freeze panes: fix header rows and pcb_name column
    ws.freeze_panes = "B3"

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    wb.save(output_path)


# =============================================================================
# Main Processing
# =============================================================================

def process_files(filepaths: List[str]) -> List[Dict]:
    """Process multiple files and return list of feature records."""
    results = []
    total = len(filepaths)

    for i, filepath in enumerate(filepaths, start=1):
        filename = Path(filepath).stem
        print(f"  [{i}/{total}] {filename}...", end=" ", flush=True)

        try:
            Z = read_elevation(filepath)
            features = extract_all_features(Z)
            record = {"pcb_name": filename}
            record.update(features)
            results.append(record)
            print(f"OK (max={features['max_warpage']:.1f}, mode={features['warpage_mode']})")
        except Exception as e:
            print(f"FAILED ({e})")

    return results


def main():
    parser = argparse.ArgumentParser(
        description="PCB Warpage Parameter Extractor"
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--input-file", type=str,
                       help="Single .txt file to process.")
    group.add_argument("--input-dir", type=str,
                       help="Directory containing .txt files to process.")
    parser.add_argument("--output", type=str, default=None,
                        help="Output Excel file path (default: warpage_features.xlsx in input location).")

    args = parser.parse_args()

    start_time = time.time()
    print("\n" + "=" * 60)
    print("  PCB Warpage Parameter Extractor")
    print("=" * 60)

    # Discover files
    if args.input_file:
        if not os.path.isfile(args.input_file):
            print(f"  [ERROR] File not found: {args.input_file}")
            sys.exit(1)
        filepaths = [args.input_file]
        default_output_dir = os.path.dirname(args.input_file) or "."
        print(f"  Mode: Single file")
    else:
        if not os.path.isdir(args.input_dir):
            print(f"  [ERROR] Directory not found: {args.input_dir}")
            sys.exit(1)
        filepaths = discover_txt_files(args.input_dir)
        default_output_dir = args.input_dir
        print(f"  Mode: Directory ({len(filepaths)} .txt files found)")

    if not filepaths:
        print("  No .txt files found. Nothing to do.")
        sys.exit(0)

    # Process
    print(f"\n  Processing {len(filepaths)} file(s)...\n")
    results = process_files(filepaths)

    if not results:
        print("\n  No files were successfully processed.")
        sys.exit(1)

    # Export
    output_path = args.output or os.path.join(default_output_dir, "warpage_features.xlsx")
    print(f"\n  Exporting to: {output_path}")
    export_to_excel(results, output_path)

    elapsed = time.time() - start_time
    print(f"\n{'=' * 60}")
    print(f"  {len(results)} file(s) processed, results saved to {output_path}")
    print(f"  Completed in {elapsed:.1f} seconds.")
    print(f"{'=' * 60}\n")


if __name__ == "__main__":
    main()
